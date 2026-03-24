"""
Отчёт по заявкам для админ-панели.
Источник: реестр привязок issue_binding_registry + Jira details.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import aiohttp

from config import CONFIG
from core.jira_aa import get_issue_admin_details
from core.jira_status_ru import jira_status_display_ru
from core.support.api import get_jira_browse_url
from core.support.issue_binding_registry import get_all_bindings
from user_storage import get_user_profile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

REPORT_FILE = Path(__file__).resolve().parents[1] / "data" / "admin_tickets_report.xlsx"


def get_total_created_tickets_count() -> int:
    """Количество уникальных заявок, созданных через бота."""
    keys = {
        (b.get("issue_key") or "").strip().upper()
        for b in get_all_bindings()
        if (b.get("issue_key") or "").strip()
    }
    return len(keys)


def _bindings_index_by_issue() -> Dict[str, Dict[str, Any]]:
    """
    По issue_key возвращает "первую" запись привязки (по created_at),
    которую считаем автором заявки в отчёте.
    """
    out: Dict[str, Dict[str, Any]] = {}
    for b in get_all_bindings():
        key = (b.get("issue_key") or "").strip().upper()
        if not key:
            continue
        old = out.get(key)
        if old is None:
            out[key] = b
            continue
        try:
            old_ts = float(old.get("created_at") or 0)
        except Exception:
            old_ts = 0.0
        try:
            new_ts = float(b.get("created_at") or 0)
        except Exception:
            new_ts = 0.0
        if 0 < new_ts < old_ts or old_ts <= 0:
            out[key] = b
    return out


def _author_from_binding(binding: Dict[str, Any]) -> str:
    ch = (binding.get("channel_id") or "telegram").strip().lower()
    try:
        uid = int(binding.get("channel_user_id"))
    except Exception:
        uid = 0
    profile = get_user_profile(uid, ch) or {}
    full_name = (profile.get("full_name") or "").strip()
    login = (profile.get("login") or "").strip()
    if full_name and login:
        return f"{full_name} ({login})"
    return full_name or login or f"{ch}:{uid}"


def _extract_sla_value(issue_fields: Dict[str, Any], field_name_target: str) -> str:
    """
    Достаёт SLA из поля с именем "Время решения IT-услуги".
    Jira может возвращать разные структуры SLA, поэтому тут мягкий парсинг.
    """
    names = issue_fields.get("_field_names") or {}
    value = None
    for field_id, field_name in names.items():
        if (field_name or "").strip().lower() == field_name_target.lower():
            value = issue_fields.get(field_id)
            break
    if value is None:
        return "-"

    if isinstance(value, str):
        return value.strip() or "-"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, dict):
        friendly = (value.get("friendly") or "").strip()
        if friendly:
            return friendly
        ongoing = value.get("ongoingCycle") or {}
        elapsed = (ongoing.get("elapsedTime") or {}).get("friendly")
        if elapsed:
            return str(elapsed).strip()
        completed = value.get("completedCycles") or []
        if isinstance(completed, list) and completed:
            last = completed[-1] or {}
            breached = last.get("breached")
            elapsed_last = ((last.get("elapsedTime") or {}).get("friendly") or "").strip()
            if elapsed_last:
                return f"{elapsed_last}{' (breached)' if breached else ''}"
    return "-"


async def _get_issue_fields_with_names(issue_key: str) -> Optional[Dict[str, Any]]:
    jira = CONFIG.get("JIRA", {})
    base_url = (jira.get("LOGIN_URL") or "").strip().rstrip("/")
    token = (jira.get("TOKEN") or "").strip()
    if not base_url or not token:
        return None
    url = f"{base_url}/rest/api/2/issue/{issue_key}?fields=*all&expand=names"
    headers = {"Accept": "application/json", "Authorization": f"Bearer {token}"}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=20)) as resp:
                if resp.status != 200:
                    return None
                data = await resp.json()
                fields = data.get("fields") or {}
                names = data.get("names") or {}
                if isinstance(fields, dict):
                    fields["_field_names"] = names if isinstance(names, dict) else {}
                return fields if isinstance(fields, dict) else None
    except Exception:
        return None


async def build_admin_detailed_report() -> Optional[Path]:
    """
    Формирует Excel-отчёт:
    Автор, Исполнитель, Номер заявки, Статус, Ссылка в Jira, SLA "Время решения IT-услуги".
    """
    if not HAS_OPENPYXL:
        return None

    by_issue = _bindings_index_by_issue()
    issue_keys = sorted(by_issue.keys())
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    headers = [
        "Автор",
        "Исполнитель",
        "Номер заявки",
        "Статус",
        "Ссылка в JIRA",
        'SL "Время решения IT-услуги"',
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for issue_key in issue_keys:
        binding = by_issue[issue_key]
        author = _author_from_binding(binding)
        info = await get_issue_admin_details(issue_key) or {}
        assignee = (info.get("assignee_display") or "—").strip() or "—"
        status = jira_status_display_ru(info.get("status"))
        jira_link = get_jira_browse_url(issue_key) or ""

        fields = await _get_issue_fields_with_names(issue_key)
        sla_value = _extract_sla_value(fields or {}, "Время решения IT-услуги")

        ws.cell(row=row, column=1, value=author)
        ws.cell(row=row, column=2, value=assignee)
        ws.cell(row=row, column=3, value=issue_key)
        ws.cell(row=row, column=4, value=status)
        ws.cell(row=row, column=5, value=jira_link)
        ws.cell(row=row, column=6, value=sla_value)
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 56
    ws.column_dimensions["F"].width = 28

    wb.save(str(REPORT_FILE))
    return REPORT_FILE
